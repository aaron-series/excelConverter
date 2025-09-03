"""
Excel Parser Module

Excel 파일(.xlsx, .xls)을 파싱하여 데이터와 스타일 정보를 추출하는 모듈
openpyxl을 사용하여 셀 데이터, 스타일, 병합 셀 정보 등을 처리합니다.
"""

from typing import Dict, List, Tuple, Optional, Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class ExcelParser:
    """Excel 파일을 파싱하여 데이터와 스타일 정보를 추출하는 클래스"""
    
    def __init__(self, file_path: str):
        """
        ExcelParser 초기화
        
        Args:
            file_path (str): Excel 파일 경로
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        
    def load_workbook(self) -> bool:
        """
        Excel 워크북을 로드합니다.
        
        Returns:
            bool: 로드 성공 여부
        """
        try:
            self.workbook = load_workbook(self.file_path, data_only=False)
            logger.info(f"Excel 파일 로드 성공: {self.file_path}")
            return True
        except FileNotFoundError:
            logger.error(f"Excel 파일을 찾을 수 없습니다: {self.file_path}")
            return False
        except Exception as e:
            logger.error(f"Excel 파일 로드 실패: {str(e)}")
            return False
    
    def get_sheet_names(self) -> List[str]:
        """
        워크북의 모든 시트 이름을 반환합니다.
        
        Returns:
            List[str]: 시트 이름 리스트
        """
        if not self.workbook:
            return []
        return self.workbook.sheetnames
    
    def select_sheet(self, sheet_name: Optional[str] = None, sheet_index: Optional[int] = None) -> bool:
        """
        특정 시트를 선택합니다.
        
        Args:
            sheet_name (str, optional): 시트 이름
            sheet_index (int, optional): 시트 인덱스
            
        Returns:
            bool: 선택 성공 여부
        """
        if not self.workbook:
            return False
            
        try:
            if sheet_name:
                self.worksheet = self.workbook[sheet_name]
            elif sheet_index is not None:
                self.worksheet = self.workbook.worksheets[sheet_index]
            else:
                # 기본값: 첫 번째 시트
                self.worksheet = self.workbook.active
                
            logger.info(f"시트 선택: {self.worksheet.title}")
            return True
        except Exception as e:
            logger.error(f"시트 선택 실패: {str(e)}")
            return False
    
    def get_used_range(self) -> Tuple[str, str]:
        """
        사용된 범위를 반환합니다.
        
        Returns:
            Tuple[str, str]: (시작 셀, 끝 셀) 예: ('A1', 'D10')
        """
        if not self.worksheet:
            return ('A1', 'A1')
            
        return (self.worksheet.min_row, self.worksheet.max_row,
                self.worksheet.min_column, self.worksheet.max_column)
    
    def extract_cell_data(self, row: int, col: int) -> Dict[str, Any]:
        """
        특정 셀의 데이터와 스타일 정보를 추출합니다.
        
        Args:
            row (int): 행 번호
            col (int): 열 번호
            
        Returns:
            Dict[str, Any]: 셀 데이터와 스타일 정보
        """
        if not self.worksheet:
            return {}
            
        cell = self.worksheet.cell(row=row, column=col)
        
        # 셀 값 처리 - 수식이면 계산된 값 사용
        cell_value = cell.value
        
        # 수식 처리
        if cell_value and str(cell_value).startswith('='):
            try:
                # 수식의 계산된 값 가져오기
                calculated_value = cell.value
                # data_only=True로 워크북을 다시 로드하여 계산된 값 추출
                temp_workbook = load_workbook(self.file_path, data_only=True)
                temp_worksheet = temp_workbook[self.worksheet.title]
                temp_cell = temp_worksheet.cell(row=row, column=col)
                calculated_value = temp_cell.value
                temp_workbook.close()
                
                # 계산된 값이 있으면 사용, 없으면 원본 수식 사용
                if calculated_value is not None:
                    cell_value = calculated_value
            except Exception as e:
                logger.warning(f"수식 계산 실패 ({cell.coordinate}): {str(e)}")
                # 수식 계산 실패 시 원본 값 사용
                pass
        
        # 기본 데이터
        cell_data = {
            'row': row,
            'col': col,
            'address': cell.coordinate,
            'value': cell_value,
            'data_type': type(cell_value).__name__,
            'number_format': cell.number_format,
            'is_merged': False,
            'merge_range': None,
            'original_value': cell.value,  # 원본 값 보존
            'is_formula': str(cell.value).startswith('=') if cell.value else False
        }
        
        # 스타일 정보 추출
        cell_data.update(self._extract_cell_style(cell))
        
        return cell_data
    
    def _extract_cell_style(self, cell) -> Dict[str, Any]:
        """
        셀의 스타일 정보를 추출합니다.
        
        Args:
            cell: openpyxl Cell 객체
            
        Returns:
            Dict[str, Any]: 스타일 정보
        """
        style_data = {
            'font': {},
            'fill': {},
            'border': {},
            'alignment': {}
        }
        
        # 폰트 정보
        if cell.font:
            font = cell.font
            style_data['font'] = {
                'name': font.name,
                'size': font.size,
                'bold': font.bold,
                'italic': font.italic,
                'underline': font.underline,
                'color': self._extract_color(font.color)
            }
        
        # 배경색 정보
        if cell.fill:
            fill = cell.fill
            style_data['fill'] = {
                'type': fill.fill_type,
                'color': self._extract_color(fill.start_color),
                'pattern_type': fill.patternType
            }
        
        # 테두리 정보
        if cell.border:
            border = cell.border
            style_data['border'] = {
                'left': self._extract_border_side(border.left),
                'right': self._extract_border_side(border.right),
                'top': self._extract_border_side(border.top),
                'bottom': self._extract_border_side(border.bottom)
            }
        
        # 정렬 정보
        if cell.alignment:
            alignment = cell.alignment
            style_data['alignment'] = {
                'horizontal': alignment.horizontal,
                'vertical': alignment.vertical,
                'wrap_text': alignment.wrap_text,
                'text_rotation': alignment.text_rotation
            }
        
        return style_data
    
    def _extract_color(self, color) -> Optional[str]:
        """
        색상 정보를 추출합니다.
        
        Args:
            color: openpyxl Color 객체
            
        Returns:
            Optional[str]: 색상 값 (RGB 또는 테마 색상)
        """
        if not color:
            return None
        
        try:
            if hasattr(color, 'rgb') and color.rgb:
                # RGB 값이 있으면 반환 (예: "FF0000")
                return str(color.rgb)
            elif hasattr(color, 'theme') and color.theme is not None:
                return f"theme_{color.theme}"
            elif hasattr(color, 'indexed') and color.indexed is not None:
                return f"indexed_{color.indexed}"
            elif hasattr(color, 'type') and color.type == 'rgb':
                # RGB 타입인 경우 기본 색상 반환
                return "000000"
            elif hasattr(color, '__str__'):
                # 문자열로 변환 가능한 경우
                color_str = str(color)
                if color_str and color_str != 'None':
                    return color_str
            
            return None
        except Exception:
            # 예외 발생 시 None 반환
            return None
    
    def _extract_border_side(self, border_side) -> Dict[str, Any]:
        """
        테두리 한쪽 면의 정보를 추출합니다.
        
        Args:
            border_side: openpyxl BorderSide 객체
            
        Returns:
            Dict[str, Any]: 테두리 정보
        """
        if not border_side:
            return {'style': None, 'color': None}
            
        return {
            'style': border_side.style,
            'color': self._extract_color(border_side.color)
        }
    
    def extract_merged_cells(self) -> List[Dict[str, Any]]:
        """
        병합된 셀 정보를 추출합니다.
        
        Returns:
            List[Dict[str, Any]]: 병합 셀 정보 리스트
        """
        if not self.worksheet:
            return []
            
        merged_cells = []
        for merged_range in self.worksheet.merged_cells.ranges:
            try:
                # 시작 셀과 끝 셀 주소 계산
                start_address = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
                end_address = f"{get_column_letter(merged_range.max_col)}{merged_range.max_row}"
                
                merged_cells.append({
                    'range': str(merged_range),
                    'start_row': merged_range.min_row,
                    'end_row': merged_range.max_row,
                    'start_col': merged_range.min_col,
                    'end_col': merged_range.max_col,
                    'start_address': start_address,
                    'end_address': end_address
                })
            except Exception as e:
                logger.warning(f"병합 셀 정보 추출 실패: {str(e)}")
                continue
        
        return merged_cells
    
    def extract_dimensions(self) -> Dict[str, Any]:
        """
        행 높이와 열 너비 정보를 추출합니다.
        
        Returns:
            Dict[str, Any]: 차원 정보
        """
        if not self.worksheet:
            return {}
            
        dimensions = {
            'row_heights': {},
            'column_widths': {},
            'default_row_height': self.worksheet.sheet_format.defaultRowHeight,
            'default_column_width': self.worksheet.sheet_format.defaultColWidth
        }
        
        # 행 높이
        for row in range(1, self.worksheet.max_row + 1):
            height = self.worksheet.row_dimensions[row].height
            if height:
                dimensions['row_heights'][row] = height
        
        # 열 너비
        for col in range(1, self.worksheet.max_column + 1):
            width = self.worksheet.column_dimensions[get_column_letter(col)].width
            if width:
                dimensions['column_widths'][col] = width
        
        return dimensions
    
    def extract_sheet_data(self, range_start: Optional[str] = None, range_end: Optional[str] = None) -> Dict[str, Any]:
        """
        시트의 모든 데이터를 추출합니다.
        
        Args:
            range_start (str, optional): 시작 범위 (예: 'A1')
            range_end (str, optional): 끝 범위 (예: 'D10')
            
        Returns:
            Dict[str, Any]: 시트 데이터
        """
        if not self.worksheet:
            return {}
        
        # 범위 결정
        if range_start and range_end:
            # 사용자 지정 범위
            start_cell = self.worksheet[range_start]
            end_cell = self.worksheet[range_end]
            min_row, max_row = start_cell.row, end_cell.row
            min_col, max_col = start_cell.column, end_cell.column
        else:
            # 전체 사용 범위
            min_row, max_row = self.worksheet.min_row, self.worksheet.max_row
            min_col, max_col = self.worksheet.min_column, self.worksheet.max_column
        
        sheet_data = {
            'sheet_name': self.worksheet.title,
            'range': f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
            'dimensions': {
                'rows': max_row - min_row + 1,
                'columns': max_col - min_col + 1,
                'start_row': min_row,
                'end_row': max_row,
                'start_col': min_col,
                'end_col': max_col
            },
            'cells': [],
            'merged_cells': self.extract_merged_cells(),
            'row_heights': self.extract_dimensions()['row_heights'],
            'column_widths': self.extract_dimensions()['column_widths']
        }
        
        # 셀 데이터 추출
        for row in range(min_row, max_row + 1):
            row_data = []
            for col in range(min_col, max_col + 1):
                cell_data = self.extract_cell_data(row, col)
                
                # 병합 셀 정보 추가
                for merged_cell in sheet_data['merged_cells']:
                    if (merged_cell['start_row'] <= row <= merged_cell['end_row'] and
                        merged_cell['start_col'] <= col <= merged_cell['end_col']):
                        cell_data['is_merged'] = True
                        cell_data['merge_range'] = merged_cell['range']
                        break
                
                row_data.append(cell_data)
            sheet_data['cells'].append(row_data)
        
        return sheet_data
    
    def close(self):
        """워크북을 닫습니다."""
        if self.workbook:
            self.workbook.close()
            logger.info("Excel 워크북을 닫았습니다.")
    
    def parse_sheet(self, sheet_name: Optional[str] = None, sheet_index: Optional[int] = None,
                   range_start: Optional[str] = None, range_end: Optional[str] = None) -> Dict[str, Any]:
        """
        시트를 파싱하는 편의 메서드
        
        Args:
            sheet_name (str, optional): 시트 이름
            sheet_index (int, optional): 시트 인덱스
            range_start (str, optional): 시작 범위
            range_end (str, optional): 끝 범위
            
        Returns:
            Dict[str, Any]: 파싱된 시트 데이터
        """
        if not self.load_workbook():
            raise Exception("Excel 파일을 로드할 수 없습니다.")
        
        if not self.select_sheet(sheet_name=sheet_name, sheet_index=sheet_index):
            raise Exception("지정된 시트를 찾을 수 없습니다.")
        
        return self.extract_sheet_data(range_start, range_end)


def parse_excel_file(file_path: str, sheet_name: Optional[str] = None, 
                    range_start: Optional[str] = None, range_end: Optional[str] = None) -> Dict[str, Any]:
    """
    Excel 파일을 파싱하는 편의 함수
    
    Args:
        file_path (str): Excel 파일 경로
        sheet_name (str, optional): 시트 이름
        range_start (str, optional): 시작 범위
        range_end (str, optional): 끝 범위
        
    Returns:
        Dict[str, Any]: 파싱된 데이터
    """
    parser = ExcelParser(file_path)
    
    try:
        if not parser.load_workbook():
            return {}
        
        if not parser.select_sheet(sheet_name):
            return {}
        
        return parser.extract_sheet_data(range_start, range_end)
    
    finally:
        parser.close()
